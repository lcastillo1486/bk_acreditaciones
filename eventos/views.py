from django.shortcuts import render, redirect
from .forms import formEvento
from .models import bkt_eventos, acreditados_tmp, acreditados_def, acreditadorEvento, inventarioBrazalete, inventarioBrazaleteAcreditardor
from django.contrib import messages
import pandas as pd
from django.http import FileResponse
from django.conf import settings
import os 
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib.auth.models import User
from datetime import date
from user_agents import parse
from django.utils import timezone
from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from collections import defaultdict
import matplotlib.pyplot as ptl
from io import BytesIO
import base64
from django.db.models import Count
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, legal, portrait
from reportlab.lib.units import cm

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFont, registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import fonts
import io
from django.db.models import Count, Case, When, IntegerField
# Create your views here.

@login_required
def eventos(request):
    form_evento = formEvento()
    eventos_activos = bkt_eventos.objects.filter(evento_activo=1)
    return render(request, 'evento.html', {'form_evento': form_evento, 'listado_eventos': eventos_activos})
@login_required
def guardarEvento(request):

    fecha_hoy = date.today()

    if not request.user.is_superuser:
        messages.error(request,'¡No tiene permisos para crear un nuevo evento!')
        return redirect('evento')
    
    if request.method == 'POST':
        form = formEvento(request.POST)
        if form.is_valid():

            nombre_even = form.cleaned_data['nombre_evento']
            lugar_even = form.cleaned_data['lugar_evento']
            fecha_even = form.cleaned_data['fecha_evento']
            if fecha_even < fecha_hoy:
                messages.error(request,'¡No se puede guardar un evento con una fecha anterior!')
                return redirect('evento')
            guarda_evento = bkt_eventos(
                nombre_evento=nombre_even, lugar_evento=lugar_even, fecha_evento=fecha_even)
            guarda_evento.save()

            messages.success(
                request, '¡Los datos se han guardado correctamente!')
            return redirect('evento')
        else:
            messages.success(request, '¡No se han guardado los datos!')
            return redirect('evento')
@login_required
def importarExcel(request, id_evento):

    if not request.user.is_superuser:
        messages.error(request,'¡No tiene permisos para importar listados!')
        return redirect('evento')

    if request.method == 'POST' and 'archivo_excel' in request.FILES:
        archivo = request.FILES['archivo_excel']

        # Leer el archivo Excel utilizando pandas
        try:
            df = pd.read_excel(archivo, dtype={'NUMERO_DOCUMENTO': str})
        except:
            # Manejar el error si el archivo no se puede leer correctamente
            messages.error(request, '¡El archivo no es válido!')
            return redirect('evento')

        # verifica que este iniciado el proceso de actualización

        iniciado_act = bkt_eventos.objects.get(id = id_evento)
        if iniciado_act.acreditacion_activa == 1:
            messages.error(request, '¡Ya se ha iniciado el proceso de acreditación!')
            return redirect('evento')


        # Verificar que no se haya alcanzado el numero maximo de actualizaciones

        num_maximo = bkt_eventos.objects.get(id=id_evento)
        maximo = num_maximo.cargas_max
        cargas_real = num_maximo.num_cargas

        # if maximo == cargas_real:
        #     messages.error(
        #         request, '¡Ha alcanzado el número máximo de cargas permitido! \n Comuníquese con los encargados del sistema si requiere una nueva actualización.')
        #     return redirect('evento')

        # Verificar las columnas requeridas
        columnas_requeridas = ['NOMBRES', 'APELLIDOS', 'TIPO_DOCUMENTO',
                               'NUMERO_DOCUMENTO', 'CARGO Y O FUNCIÓN', 'AREA_DE_TRABAJO', 'COLOR_ZONA_BRAZALETE']
        columnas_excel = df.columns.tolist()
        if not set(columnas_requeridas).issubset(columnas_excel):
            # Manejar el error si alguna(s) columna(s) requerida(s) no está presente
            messages.error(
                request, '¡El archivo no contiene todas las columnas requeridas!')
            return redirect('evento')

        # Validar campos vacíos
        registros = []
        for _, row in df.iterrows():
            if any(pd.isnull(row[columna]) or str(row[columna]).strip() == '' for columna in columnas_requeridas):
                # Manejar el error si hay campos vacíos en las columnas requeridas
                messages.error(
                    request, '¡El archivo contiene campos vacíos en las columnas requeridas! Por favor corrija e intente nuevamente.')
                return redirect('evento')

        # Procesar los datos y guardar en la base de datos
        registros = []

        for  _, row in df.iterrows():
            registro = acreditados_tmp(
                nombre_persona=row['NOMBRES'],
                apellido_persona=row['APELLIDOS'],
                tipo_doc=row['TIPO_DOCUMENTO'],
                numero_doc=row['NUMERO_DOCUMENTO'].strip().replace(" ", ""),
                cargo=row['CARGO Y O FUNCIÓN'],
                zona_acceso=row['AREA_DE_TRABAJO'],
                color_zona = row['COLOR_ZONA_BRAZALETE'],
                empresa=row['EMPRESA'],
                id_evento_id = id_evento
            )
            registros.append(registro)

        ### Borrar la tabla si tiene registros ###

        if acreditados_tmp.objects.filter(id_evento_id = id_evento).exists():
            acreditados_tmp.objects.filter(id_evento_id = id_evento).delete()

        acreditados_tmp.objects.bulk_create(registros)

        actualiza_cargas = bkt_eventos.objects.get(id=id_evento)
        actualiza_cargas.num_cargas = actualiza_cargas.num_cargas + 1
        actualiza_cargas.save()

        messages.success(request, '¡Los datos se han importado exitosamente!')
        return redirect('evento')

    return redirect('evento')
@login_required
def iniciaAcreditacion(request, id_evento):

    if not request.user.is_superuser:
        messages.error(request,'¡No tiene permisos para iniciar el proceso de acrecitación!')
        return redirect('evento')
    
    # verifica si se subio listado de brazaletes

    # if not inventarioBrazalete.objects.filter(id_evento = id_evento, evento_cerrado = 0).exists():
    #     messages.error(request, '¡Debe importar el inventario de brazaletes para poder iniciar el proceso de acreditación!')
    #     return redirect('evento')
    
    # verifica si ya esta activada 

    if bkt_eventos.objects.filter(id = id_evento, acreditacion_activa = 1).exists():
        messages.error(request, '¡El proceso de acreditación ya se ha iniciado!')
        return redirect('evento')
    else:
        tabla_origen = acreditados_tmp.objects.filter(id_evento_id = id_evento)

        tabla_destino =[]  
            
        for dato in tabla_origen:
            registros = acreditados_def(
            nombre_persona = dato.nombre_persona,
            apellido_persona = dato.apellido_persona,
            tipo_doc = dato.tipo_doc,
            numero_doc = dato.numero_doc,
            cargo = dato.cargo,
            zona_acceso = dato.zona_acceso,
            color_zona = dato.color_zona,
            empresa = dato.empresa,
            id_evento_id = id_evento)

            tabla_destino.append(registros)

        acreditados_def.objects.bulk_create(tabla_destino)

        
        iniciar_bkt = bkt_eventos.objects.get(id = id_evento)
        iniciar_bkt.acreditacion_activa = 1
        iniciar_bkt.save()

        acreditados_tmp.objects.filter(id_evento_id=id_evento).delete()
        return redirect('evento')
@login_required
def detieneAcreditacion(request, id_evento):

    if not request.user.is_superuser:
        messages.error(request,'¡No tiene permisos para detener el proceso de acrecitación!')
        return redirect('evento')
    
    desactiva_bkt = bkt_eventos.objects.get(id = id_evento)
    if desactiva_bkt.acreditacion_activa == 0:
        messages.error(request, '¡No se ha iniciado el proceso de acreditación!')
        return redirect('evento')
    else:
        desactiva_bkt.acreditacion_activa = 0
        desactiva_bkt.evento_activo = 0
        desactiva_bkt.save()

        # desestima el listado definitivo
        acreditados_def.objects.filter(id_evento_id = id_evento).update(evento_cerrado=1)
        acreditadorEvento.objects.filter(evento = id_evento).update(cerrado = 1)
        inventarioBrazalete.objects.filter(id_evento = id_evento).update(evento_cerrado = 1)
        inventarioBrazaleteAcreditardor.objects.filter(id_evento = id_evento).update(evento_cerrado = 1)


        return redirect('evento')
@login_required
def descargaFormato(request):

    ruta_archivo_zip = os.path.join(settings.BASE_DIR, 'Formato_Acreditaciones.zip')

    # Generar la respuesta con el contenido del archivo ZIP
    with open(ruta_archivo_zip, 'rb') as archivo_zip:
        contenido_zip = archivo_zip.read()

    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="Formato_Acreditaciones.zip"'
    response.write(contenido_zip)

    return response
@login_required
def verMonitor(request):

    cantidad_registros = acreditados_def.objects.filter().count()
    cantidad_acreditados = acreditados_def.objects.filter(acreditado = 1).count()
    porcentaje = cantidad_acreditados/cantidad_registros

    return render(request, 'monitor.html')
@login_required
def acreditarPersonal(request, id_reg):

    hora_actual = datetime.now()

    # Definir la diferencia horaria entre Oregon y Lima (Perú)
    diferencia_horaria = timedelta(hours=+2)  # UTC-8 (Oregon) - UTC-5 (Lima) = -3 horas

    # Calcular la hora actual en Lima (Perú)
    # hora_actual_peru = hora_actual_utc + diferencia_horaria

    # Puedes formatear la hora según tus necesidades
    formato_hora_peru = hora_actual.strftime('%H:%M')

    user_agent_string = request.META['HTTP_USER_AGENT']
    user_agent = parse(user_agent_string)

    is_mobile = user_agent.is_mobile
    is_tablet = user_agent.is_tablet

    if not bkt_eventos.objects.filter(acreditacion_activa = 1).exists():
        messages.error(request, '¡No se ha iniciado el proceso de acreditación!')
        return redirect('buscar_personal') 

    if id_reg == "":
        if is_mobile or is_tablet:
            return redirect('buscar_personal_movil')
        else:
            return redirect('buscar_personal')

    if acreditados_def.objects.filter(id = id_reg, acreditado = 1).exists():
        #busca estadisticas
        busca_stad = acreditados_def.objects.get(id = id_reg, acreditado = 1)
        id_even = busca_stad.id_evento
        total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
        total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
        porcentaje = round((total_acreditado /total_registros)*100,4)

        messages.error(request, '¡Ya fue acreditado anteriormente!')
        if is_mobile or is_tablet:
            return redirect('buscar_personal_movil')
        else:
            return render(request,'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
        
    acreditado = acreditados_def.objects.get(id = id_reg, acreditado = 0)
    acreditado.acreditado = 1
    acreditado.acreditado_por = request.user.username
    acreditado.asistencia = 1
    acreditado.hora = formato_hora_peru
    cod_evento = acreditado.id_evento_id
    acreditado.save()

    # actualizar inventario brazaletes 
    # zona = acreditado.zona_acceso
    # actu_brazalete = inventarioBrazalete.objects.get(id_evento = cod_evento, nombre_brazalete__icontains = zona)
    # actu_brazalete.cantidad_entregada = actu_brazalete.cantidad_entregada +1
    # actu_brazalete.cantidad_resta = actu_brazalete.cantidad_brazalete - actu_brazalete.cantidad_entregada
    # actu_brazalete.save()

    # actualizar inventario brazaletes acreditador
    zona = acreditado.zona_acceso
    acreditador = request.user.username
    # if inventarioBrazaleteAcreditardor.objects.filter(id_evento = cod_evento, nombre_brazalete__icontains = zona, nombre_acreditador = acreditador).exists():
    #     actu_brazalete_acred = inventarioBrazaleteAcreditardor.objects.get(id_evento = cod_evento, nombre_brazalete__icontains = zona, nombre_acreditador = acreditador)
    #     actu_brazalete_acred.cantidad_entregada = actu_brazalete_acred.cantidad_entregada +1
    #     actu_brazalete_acred.cantidad_resta = actu_brazalete_acred.cantidad_brazalete - actu_brazalete_acred.cantidad_entregada
    #     actu_brazalete_acred.save()


    #busca estadisticas
    busca_stad = acreditados_def.objects.get(id = id_reg, acreditado = 1)
    id_even = busca_stad.id_evento
    total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
    total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
    porcentaje = round((total_acreditado /total_registros)*100,4)

    messages.success(request, '¡Acreditado Correctamente!')
    if is_mobile or is_tablet:
        return redirect('buscar_personal_movil')
    else:
        return render(request,'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
@login_required
def buscarPersona(request):

    user_agent_string = request.META['HTTP_USER_AGENT']
    user_agent = parse(user_agent_string)

    is_mobile = user_agent.is_mobile
    is_tablet = user_agent.is_tablet

    if is_mobile or is_tablet:
        return redirect('buscar_personal_movil')



    usuario = request.user
    if not acreditadorEvento.objects.filter(usuario = usuario, cerrado = 0).exists():
        messages.error(request, '¡No se ha iniciado el proceso de acreditación!')
        return redirect('evento')
    else:
        evento_buscar = acreditadorEvento.objects.get(usuario = usuario, cerrado = 0)
        cod_event = evento_buscar.evento


        
    if not bkt_eventos.objects.filter(acreditacion_activa = 1).exists():
        messages.error(request, '¡No se ha iniciado el proceso de acreditación!')
        return redirect('evento')
    else:
        if bkt_eventos.objects.filter(acreditacion_activa = 1).count() > 1:
            event_act = bkt_eventos.objects.filter(acreditacion_activa = 1).count()
            cod = bkt_eventos.objects.filter(acreditacion_activa = 1)
            valores = cod.values_list('id', 'nombre_evento')
     
    if request.method == 'POST':
        #buscar por documento
        documento = request.POST.get('documento').strip()
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        nombre_empresa = request.POST.get('empresa')

        #evalua el len del documento

        if len(documento) >0 and  len(documento) < 7:
            messages.error(request, '¡El número de documento es demasiado corto!')
            return render(request, 'acredpersonal.html')

        #evaluar si los tres estan vacios
        if len(documento) == 0 and len(nombre) == 0 and len(apellido)==0 and len(nombre_empresa) ==0:
            messages.error(request, '¡Debe ingresar datos para la búsqueda!')
            return render(request, 'acredpersonal.html')
        
        #evaluar si se ha introducido solo un nombre
        if len(documento) == 0 and len(nombre) > 0 and len(apellido) == 0:
            messages.error(request, '¡Para una búsqueda más precisa, introduzca támbien un apellido!')
            return render(request, 'acredpersonal.html')
        
        #evaluar si se ha introducido solo el apellido
        if len(documento) == 0 and len(nombre) == 0 and len(apellido) > 0:
            messages.error(request, '¡Para una búsqueda más precisa, introduzca támbien un nombre!')
            return render(request, 'acredpersonal.html')
        
        #si busca empresa solamente
        if len(documento) == 0 and len(nombre) == 0 and len(apellido) == 0 and len(nombre_empresa) > 0:
            if acreditados_def.objects.filter(empresa__icontains = nombre_empresa, id_evento_id = cod_event).exists():    
                personal_empresa = acreditados_def.objects.filter(empresa__icontains = nombre_empresa, id_evento_id = cod_event).order_by('apellido_persona')
                return render(request,'personalempresa.html', {'personal':personal_empresa})
            else:
                messages.error( request,'¡La empresa indicada no existe en los registros de este evento!')
                total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)
                return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
        
        #busca por nombre y apellido
        if len(documento) == 0 and len(nombre) > 0 and len(apellido) > 0:
            if acreditados_def.objects.filter(Q(nombre_persona__icontains=nombre) & Q(apellido_persona__icontains=apellido, id_evento_id = cod_event)).exists():
                try:
                    persona = acreditados_def.objects.get(Q(nombre_persona__icontains=nombre) & Q(apellido_persona__icontains=apellido, id_evento_id = cod_event))
                    if persona.acreditado == 1:
                        nombre = persona.nombre_persona
                        apellido = persona.apellido_persona
                        documento = persona.numero_doc
                        empresa = persona.empresa
                        area = persona.zona_acceso
                        color = persona.color_zona
                        id_reg = persona.id
                        id_even = persona.id_evento_id
                        hora = persona.hora
                        acreditador = persona.acreditado_por

                        #busca nombre evento
                        nombre_event = bkt_eventos.objects.get(id = id_even)
                        event_name = nombre_event.nombre_evento

                        #busca estadisticas
                        total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                        total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                        porcentaje = round((total_acreditado /total_registros)*100,4)
                        messages.error(request, f'¡Ya fue acreditado anteriormente a las: {hora} por {acreditador}!')
                        return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                                    'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
                    else:
                        nombre = persona.nombre_persona
                        apellido = persona.apellido_persona
                        documento = persona.numero_doc
                        empresa = persona.empresa
                        area = persona.zona_acceso
                        color = persona.color_zona
                        id_reg = persona.id
                        id_even = persona.id_evento_id

                        #busca nombre evento
                        nombre_event = bkt_eventos.objects.get(id = id_even)
                        event_name = nombre_event.nombre_evento

                        #busca estadisticas
                        total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                        total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                        porcentaje = round((total_acreditado /total_registros)*100,4)

                        return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                                    'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})

                except:
                    messages.error(request,'¡Múltiples registros coinciden con los parámetros indicados, por favor realice una búsqueda número de documento o agregue un segundo apellido!')
                    return redirect('buscar_personal')
            else:
                messages.error(request, '¡No hay concidencias en la busqueda!')
                return redirect('buscar_personal')
        
        #busca por apellido
        # if len(documento) == 0 and len(nombre) == 0 and len(apellido) > 0:
        #     if acreditados_def.objects.filter(Q(apellido_persona__icontains=apellido)).exists():
        #         persona = acreditados_def.objects.get(Q(apellido_persona__icontains=apellido),asistencia = 1, evento_cerrado = 0)
        #         if persona.acreditado == 1:
        #             messages.error(request, '¡Esta persona ya fue acreditada anteriormente!')
        #             nombre = persona.nombre_persona
        #             apellido = persona.apellido_persona
        #             documento = persona.numero_doc
        #             empresa = persona.empresa
        #             area = persona.zona_acceso
        #             id_reg = persona.id
        #             id_even = persona.id_evento_id

        #             #busca nombre evento
        #             nombre_event = bkt_eventos.objects.get(id = id_even)
        #             event_name = nombre_event.nombre_evento

        #             #busca estadisticas
        #             total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
        #             total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
        #             porcentaje = round((total_acreditado /total_registros)*100,4)

        #             return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area, 'id':id_reg, 'evento':event_name,
        #                                                          'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
        #         else:
        #             nombre = persona.nombre_persona
        #             apellido = persona.apellido_persona
        #             documento = persona.numero_doc
        #             empresa = persona.empresa
        #             area = persona.zona_acceso
        #             id_reg = persona.id
        #             id_even = persona.id_evento_id

        #             #busca nombre evento
        #             nombre_event = bkt_eventos.objects.get(id = id_even)
        #             event_name = nombre_event.nombre_evento

        #             #busca estadisticas
        #             total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
        #             total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
        #             porcentaje = round((total_acreditado /total_registros)*100,4)

        #             return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area, 'id':id_reg, 'evento':event_name,
        #                                                          'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})

        if len(documento) > 0:
            doc = str(documento)[-7:]
            cuenta_reg = acreditados_def.objects.filter(numero_doc__endswith = doc, id_evento_id = cod_event).count()
            if cuenta_reg > 1:               
                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)
                messages.error(request,'¡Extrañamente hay más de un registro en este evento, que coincide con el número de documento!\n \
                               Seguramente sea un error de tipeo en el archivo. Se recomienda realizar una busqueda por nombre y apellido o por empresa')
                return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
            
        # valida si ya se acredito
            if acreditados_def.objects.filter(numero_doc__endswith = doc, acreditado = 1, asistencia = 1, id_evento_id = cod_event).exists():
                persona = acreditados_def.objects.get(numero_doc__endswith = doc, acreditado = 1, id_evento_id = cod_event)
                nombre = persona.nombre_persona
                apellido = persona.apellido_persona
                documento = persona.numero_doc
                empresa = persona.empresa
                area = persona.zona_acceso
                color = persona.color_zona
                id_reg = persona.id
                id_even = persona.id_evento_id
                hora = persona.hora
                acreditador = persona.acreditado_por
                
                #busca nombre evento
                nombre_event = bkt_eventos.objects.get(id = id_even)
                event_name = nombre_event.nombre_evento

                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)

                messages.error(request, f'¡Ya fue acreditado anteriormente a las {hora} por {acreditador}!')
                return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                              'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
                
            
            if acreditados_def.objects.filter(numero_doc__endswith = doc, acreditado = 0, asistencia = 0, id_evento_id = cod_event).exists():
                persona = acreditados_def.objects.get(numero_doc__endswith = doc, acreditado = 0, asistencia = 0, id_evento_id = cod_event)
                nombre = persona.nombre_persona
                apellido = persona.apellido_persona
                documento = persona.numero_doc
                empresa = persona.empresa
                area = persona.zona_acceso
                color = persona.color_zona
                id_reg = persona.id
                id_even = persona.id_evento_id


                #busca nombre evento
                nombre_event = bkt_eventos.objects.get(id = id_even)
                event_name = nombre_event.nombre_evento

                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)


                return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                             'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
            else:
                
                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)
                messages.error(request, '¡No hay concidencias en la busqueda!')
                return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})

    #busca estadisticas
    total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
    total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
    porcentaje = round((total_acreditado /total_registros)*100,4)
    return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
@login_required
def registraUsuario(request, cod_event):

    user_agent_string = request.META['HTTP_USER_AGENT']
    user_agent = parse(user_agent_string)

    is_mobile = user_agent.is_mobile
    is_tablet = user_agent.is_tablet
    
    usuario = request.user

    if not acreditadorEvento.objects.filter(usuario = usuario, cerrado = 0).exists():
        acreditadorEvento.objects.create(usuario = usuario, evento = cod_event, cerrado = 0)
        evento_buscar = acreditadorEvento.objects.get(usuario = usuario, evento = cod_event, cerrado = 0)
        if is_mobile or is_tablet:
            return redirect('buscar_personal_movil')
        else:
            return redirect('buscar_personal')
        
    else:
        evento_buscar = acreditadorEvento.objects.get(usuario = usuario, cerrado = 0)
        event_cod = evento_buscar.evento
        if int(event_cod) == int(cod_event):
            if is_mobile or is_tablet:
                return redirect('buscar_personal_movil')
            else:
                return redirect('buscar_personal')
            
        else:
            name_evento = bkt_eventos.objects.get(id = event_cod)
            name = name_evento.nombre_evento
            messages.error(request,f'Su usuario ya se ha registrado como acreditador el evento {name}')
            if is_mobile or is_tablet:
                return redirect('vista_movil')
            else:
                return redirect('evento')
@login_required
def vistaMovil(request):
    form_evento = formEvento()
    eventos_activos = bkt_eventos.objects.filter(evento_activo=1)
    return render(request, 'movil.html', {'form_evento': form_evento, 'listado_eventos': eventos_activos})
@login_required
def buscarPersonaMovil(request):

    usuario = request.user
    if not acreditadorEvento.objects.filter(usuario = usuario, cerrado = 0).exists():
        messages.error(request, '¡No se ha iniciado el proceso de acreditación!')
        return redirect('evento')
    else:
        evento_buscar = acreditadorEvento.objects.get(usuario = usuario, cerrado = 0)
        cod_event = evento_buscar.evento

        
    if not bkt_eventos.objects.filter(acreditacion_activa = 1).exists():
        messages.error(request, '¡No se ha iniciado el proceso de acreditación!')
        return redirect('evento')
    else:
        if bkt_eventos.objects.filter(acreditacion_activa = 1).count() > 1:
            event_act = bkt_eventos.objects.filter(acreditacion_activa = 1).count()
            cod = bkt_eventos.objects.filter(acreditacion_activa = 1)
            valores = cod.values_list('id', 'nombre_evento')
     
    if request.method == 'POST':
        #buscar por documento
        documento = request.POST.get('documento').strip()
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        nombre_empresa = request.POST.get('empresa')

        #evalua el len del documento

        if len(documento) >0 and  len(documento) < 7:
            messages.error(request, '¡El número de documento es demasiado corto!')
            return render(request, 'acredpersonalmovil.html')

        #evaluar si los tres estan vacios
        if len(documento) == 0 and len(nombre) == 0 and len(apellido)==0 and len(nombre_empresa) ==0:
            messages.error(request, '¡Debe ingresar datos para la búsqueda!')
            return render(request, 'acredpersonalmovil.html')
        
        #evaluar si se ha introducido solo un nombre
        if len(documento) == 0 and len(nombre) > 0 and len(apellido) == 0:
            messages.error(request, '¡Para una búsqueda más precisa, introduzca támbien un apellido!')
            return render(request, 'acredpersonalmovil.html')
        
        #evaluar si se ha introducido solo el apellido
        if len(documento) == 0 and len(nombre) == 0 and len(apellido) > 0:
            messages.error(request, '¡Para una búsqueda más precisa, introduzca támbien un nombre!')
            return render(request, 'acredpersonalmovil.html')
        
        #si busca empresa solamente
        if len(documento) == 0 and len(nombre) == 0 and len(apellido) == 0 and len(nombre_empresa) > 0:
            if acreditados_def.objects.filter(empresa__icontains = nombre_empresa, id_evento_id = cod_event).exists():    
                personal_empresa = acreditados_def.objects.filter(empresa__icontains = nombre_empresa, id_evento_id = cod_event)
                return render(request,'personalempresa.html', {'personal':personal_empresa})
            else:
                messages.error( request,'¡La empresa indicada no existe en los registros de este evento!')
                total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)
                return render(request, 'acredpersonalmovil.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
        
        #busca por nombre y apellido
        if len(documento) == 0 and len(nombre) > 0 and len(apellido) > 0:
            if acreditados_def.objects.filter(Q(nombre_persona__icontains=nombre) & Q(apellido_persona__icontains=apellido, id_evento_id = cod_event)).exists():
                try:
                    persona = acreditados_def.objects.get(Q(nombre_persona__icontains=nombre) & Q(apellido_persona__icontains=apellido, id_evento_id = cod_event))
                    if persona.acreditado == 1:
                        
                        nombre = persona.nombre_persona
                        apellido = persona.apellido_persona
                        documento = persona.numero_doc
                        empresa = persona.empresa
                        area = persona.zona_acceso
                        color = persona.color_zona
                        id_reg = persona.id
                        id_even = persona.id_evento_id
                        hora = persona.hora
                        acreditador = persona.acreditado_por

                        #busca nombre evento
                        nombre_event = bkt_eventos.objects.get(id = id_even)
                        event_name = nombre_event.nombre_evento

                        #busca estadisticas
                        total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                        total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                        porcentaje = round((total_acreditado /total_registros)*100,4)

                        messages.error(request, f'¡Ya fue acreditado anteriormente a las: {hora} por {acreditador}!')
                        return render(request, 'acredpersonalmovil.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                                    'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
                    else:
                        nombre = persona.nombre_persona
                        apellido = persona.apellido_persona
                        documento = persona.numero_doc
                        empresa = persona.empresa
                        area = persona.zona_acceso
                        color = persona.color_zona
                        id_reg = persona.id
                        id_even = persona.id_evento_id

                        #busca nombre evento
                        nombre_event = bkt_eventos.objects.get(id = id_even)
                        event_name = nombre_event.nombre_evento

                        #busca estadisticas
                        total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                        total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                        porcentaje = round((total_acreditado /total_registros)*100,4)

                        return render(request, 'acredpersonalmovil.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                                    'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})

                except:
                    messages.error(request,'¡Múltiples registros coinciden con los parámetros indicados, por favor realice una búsqueda número de documento o agregue un segundo apellido!')
                    return redirect('buscar_personal_movil')
            else:
                messages.error(request, '¡No hay concidencias en la busqueda!')
                return redirect('buscar_personal_movil')
        
        #busca por apellido
        # if len(documento) == 0 and len(nombre) == 0 and len(apellido) > 0:
        #     if acreditados_def.objects.filter(Q(apellido_persona__icontains=apellido)).exists():
        #         persona = acreditados_def.objects.get(Q(apellido_persona__icontains=apellido),asistencia = 1, evento_cerrado = 0)
        #         if persona.acreditado == 1:
        #             messages.error(request, '¡Esta persona ya fue acreditada anteriormente!')
        #             nombre = persona.nombre_persona
        #             apellido = persona.apellido_persona
        #             documento = persona.numero_doc
        #             empresa = persona.empresa
        #             area = persona.zona_acceso
        #             id_reg = persona.id
        #             id_even = persona.id_evento_id

        #             #busca nombre evento
        #             nombre_event = bkt_eventos.objects.get(id = id_even)
        #             event_name = nombre_event.nombre_evento

        #             #busca estadisticas
        #             total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
        #             total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
        #             porcentaje = round((total_acreditado /total_registros)*100,4)

        #             return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area, 'id':id_reg, 'evento':event_name,
        #                                                          'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
        #         else:
        #             nombre = persona.nombre_persona
        #             apellido = persona.apellido_persona
        #             documento = persona.numero_doc
        #             empresa = persona.empresa
        #             area = persona.zona_acceso
        #             id_reg = persona.id
        #             id_even = persona.id_evento_id

        #             #busca nombre evento
        #             nombre_event = bkt_eventos.objects.get(id = id_even)
        #             event_name = nombre_event.nombre_evento

        #             #busca estadisticas
        #             total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
        #             total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
        #             porcentaje = round((total_acreditado /total_registros)*100,4)

        #             return render(request, 'acredpersonal.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area, 'id':id_reg, 'evento':event_name,
        #                                                          'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})

        if len(documento) > 0:
            doc = str(documento)[-7:]

            cuenta_reg = acreditados_def.objects.filter(numero_doc__endswith = doc, id_evento_id = cod_event).count()
            if cuenta_reg > 1:               
                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)
                messages.error(request,'¡Extrañamente hay más de un registro en este evento, que coincide con el número de documento!\n \
                               Seguramente sea un error de tipeo en el archivo. Se recomienda realizar una busqueda por nombre y apellido o por empresa')
                return render(request, 'acredpersonalmovil.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
            
        # valida si ya se acredito
            if acreditados_def.objects.filter(numero_doc__endswith = doc, acreditado = 1, asistencia = 1, id_evento_id = cod_event).exists():
                persona = acreditados_def.objects.get(numero_doc__endswith = doc, acreditado = 1, id_evento_id = cod_event)
                nombre = persona.nombre_persona
                apellido = persona.apellido_persona
                documento = persona.numero_doc
                empresa = persona.empresa
                area = persona.zona_acceso
                color = persona.color_zona
                id_reg = persona.id
                id_even = persona.id_evento_id
                hora = persona.hora
                acreditador = persona.acreditado_por
                
                #busca nombre evento
                nombre_event = bkt_eventos.objects.get(id = id_even)
                event_name = nombre_event.nombre_evento
                messages.error(request, f'¡Ya fue acreditado anteriormente a las {hora} por {acreditador}!')

                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)

                return render(request, 'acredpersonalmovil.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                              'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
                
            
            if acreditados_def.objects.filter(numero_doc__endswith = doc, acreditado = 0, asistencia = 0, id_evento_id = cod_event).exists():
                persona = acreditados_def.objects.get(numero_doc__endswith = doc, acreditado = 0, asistencia = 0, id_evento_id = cod_event)
                nombre = persona.nombre_persona
                apellido = persona.apellido_persona
                documento = persona.numero_doc
                empresa = persona.empresa
                area = persona.zona_acceso
                color = persona.color_zona
                id_reg = persona.id
                id_even = persona.id_evento_id


                #busca nombre evento
                nombre_event = bkt_eventos.objects.get(id = id_even)
                event_name = nombre_event.nombre_evento

                #busca estadisticas
                total_acreditado = acreditados_def.objects.filter(id_evento_id = id_even, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = id_even).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)


                return render(request, 'acredpersonalmovil.html',{'nombre':nombre, 'apellido':apellido, 'documento':documento, 'empresa':empresa, 'zona':area,'color':color, 'id':id_reg, 'evento':event_name,
                                                             'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
            else:
                
                messages.error(request, '¡Los datos suministrados no coinciden con ningún registro!')
                return render(request, 'acredpersonalmovil.html')

    return render(request, 'acredpersonalmovil.html')
@login_required
def acreditacionMultiple(request):

    hora_actual = datetime.now()

    # Definir la diferencia horaria entre Oregon y Lima (Perú)
    diferencia_horaria = timedelta(hours=-2)  # UTC-8 (Oregon) - UTC-5 (Lima) = -3 horas

    # Calcular la hora actual en Lima (Perú)
    # hora_actual_peru = hora_actual_utc + diferencia_horaria

    # Puedes formatear la hora según tus necesidades
    formato_hora_peru = hora_actual.strftime('%H:%M')

    user_agent_string = request.META['HTTP_USER_AGENT']
    user_agent = parse(user_agent_string)

    is_mobile = user_agent.is_mobile
    is_tablet = user_agent.is_tablet

    usuario = request.user
    evento_buscar = acreditadorEvento.objects.get(usuario = usuario, cerrado = 0)
    cod_event = evento_buscar.evento

    if request.method == 'POST':
        registros_seleccionados = request.POST.getlist('regitrosAcreditar')

        if not registros_seleccionados:
            if is_mobile or is_tablet:
                return redirect('buscar_personal_movil')
            else:
                total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
                total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
                porcentaje = round((total_acreditado /total_registros)*100,4)
                return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})


        for registros_id in registros_seleccionados:
            registro = acreditados_def.objects.get(id = registros_id)
            registro.acreditado = 1
            registro.acreditado_por = request.user.username
            registro.asistencia = 1
            registro.hora = formato_hora_peru
            registro.save()

            # actualizar inventario brazaletes 
            zona = registro.zona_acceso
            cod_evento = registro.id_evento_id
            # actu_brazalete = inventarioBrazalete.objects.get(id_evento = cod_evento, nombre_brazalete__icontains = zona)
            # actu_brazalete.cantidad_entregada = actu_brazalete.cantidad_entregada +1
            # actu_brazalete.cantidad_resta = actu_brazalete.cantidad_brazalete - actu_brazalete.cantidad_entregada
            # actu_brazalete.save()

            # actualizar inventario brazaletes acreditador
            zona = registro.zona_acceso
            acreditador = request.user.username
            if inventarioBrazaleteAcreditardor.objects.filter(id_evento = cod_evento, nombre_brazalete__icontains = zona, nombre_acreditador = acreditador).exists():
                actu_brazalete_acred = inventarioBrazaleteAcreditardor.objects.get(id_evento = cod_evento, nombre_brazalete__icontains = zona, nombre_acreditador = acreditador)
                actu_brazalete_acred.cantidad_entregada = actu_brazalete_acred.cantidad_entregada +1
                actu_brazalete_acred.cantidad_resta = actu_brazalete_acred.cantidad_brazalete - actu_brazalete_acred.cantidad_entregada
                actu_brazalete_acred.save()

        
        total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
        total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
        porcentaje = round((total_acreditado /total_registros)*100,4)
        messages.success(request, '¡Acreditados Correctamente!')
        if is_mobile or is_tablet:
            return redirect('buscar_personal_movil')
        else:
            return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
    
    if is_mobile or is_tablet:
        return redirect('buscar_personal_movil')
    else:
        total_acreditado = acreditados_def.objects.filter(id_evento_id = cod_event, acreditado = 1).count()
        total_registros = acreditados_def.objects.filter(id_evento_id = cod_event).count()
        porcentaje = round((total_acreditado /total_registros)*100,4)
        return render(request, 'acredpersonal.html',{'total_acreditado':total_acreditado, 'total_registros':total_registros, 'porcentaje':porcentaje})
@login_required
def vistaSensei(request):
    form_evento = formEvento()
    eventos_activos = bkt_eventos.objects.filter(evento_activo=1)
    return render(request, 'PanelDeLuis.html', {'form_evento': form_evento, 'listado_eventos': eventos_activos})
@login_required
def exportarExcel(request, id):
    #recibir aqui el id del evento. Necesito un litsado de eventos
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Detalle_Acreditados.xlsx'

    # Crear un libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado_General"

    # este es el Encabezados de las columnas
    ws.append(['Nombres', 'Apellidos', 'Tipo Documento','Numero Documento','Cargo/Función','Empresa', 'Zona','¿Acreditado?'])  

    queryset = acreditados_def.objects.filter(evento_cerrado=1, id_evento_id = id).order_by('apellido_persona') 
    for item in queryset:
        if item.acreditado == 1:
            item.acreditado = 'Si'
        else:
            item.acreditado = 'No'
        ws.append([item.nombre_persona, item.apellido_persona, item.tipo_doc, item.numero_doc, item.cargo,item.empresa, item.zona_acceso, item.acreditado])  
    
    #empresa
    queryset_empresa = acreditados_def.objects.filter(evento_cerrado=1, id_evento_id=id).order_by('apellido_persona')

    empleados_por_empresa = {}
    # Iterar a través de cada elemento en la queryset y agrupar por empresa
    # Iterar a través de cada elemento en la queryset y agrupar por empresa
    for item1 in queryset_empresa:
        nombre_empresa = item1.empresa


        if nombre_empresa not in empleados_por_empresa:
            empleados_por_empresa[nombre_empresa] = []

        empleados_por_empresa[nombre_empresa].append(item1)

    # Crear hojas para cada empresa y listar empleados
    for nombre_empresa, empleados in empleados_por_empresa.items():
        nueva_hoja = wb.create_sheet(title=nombre_empresa)
        nueva_hoja.append(['Nombres', 'Apellidos', 'Documento', 'Cargo/Función', 'Zona', '¿Acreditado?'])

        for empleado in empleados:
            if empleado.acreditado == 1:
                empleado.acreditado = 'Si'
            else:
                empleado.acreditado = 'No'
            nueva_hoja.append([empleado.nombre_persona, empleado.apellido_persona, empleado.numero_doc, empleado.cargo, empleado.zona_acceso, empleado.acreditado])

    ###TOTALES#####

    queryset_totales = acreditados_def.objects.filter(evento_cerrado=1, id_evento_id=id).order_by('apellido_persona')

# Crear un diccionario para agrupar acreditados por empresa y zona
    empleados_por_empresa_zona = defaultdict(list)

    # Iterar a través de cada elemento en la queryset y agrupar por empresa y zona
    for item in queryset_totales:
        nombre_empresa = item.empresa
        zona = item.zona_acceso
        empleados_por_empresa_zona[(nombre_empresa, zona)].append(item)

    # Crear una nueva hoja para los totales generales
    hoja_totales = wb.create_sheet(title="Totales_Empresa")

    hoja_totales.append(['Empresa', 'Zona', 'Total Acreditados', 'Total No Acreditados'])

    # Calcular los totales generales por empresa y zona
    for (nombre_empresa, zona), empleados in empleados_por_empresa_zona.items():
        total_acreditados = sum(1 for empleado in empleados if empleado.acreditado == 1)
        total_no_acreditados = sum(1 for empleado in empleados if empleado.acreditado != 1)

        hoja_totales.append([nombre_empresa, zona, total_acreditados, total_no_acreditados]) 

    
    ###RESUMEN BRAZALETES#####
    
    hoja_brazaletes = wb.create_sheet(title="Inventario_Brazaletes")

    queryset_brazaletes = inventarioBrazalete.objects.filter(id_evento=id, evento_cerrado = 1)
    hoja_brazaletes.append(['Tipo Brazalete', 'Cantidad Inicial', 'Cantidad Entregada', 'Inventario Final'])
    
    for item in queryset_brazaletes:
        hoja_brazaletes.append([item.nombre_brazalete, item.cantidad_brazalete, item.cantidad_entregada, item.cantidad_resta])
    
    ###RESUMEN BRAZALETES ACREDITADOR#####
    
    hoja_brazaletes_acreditador = wb.create_sheet(title="Inventario_Acreditador")

    queryset_brazaletes_acreditador = inventarioBrazaleteAcreditardor.objects.filter(id_evento=id, evento_cerrado = 1).order_by('nombre_acreditador')
    hoja_brazaletes_acreditador.append(['Acreditador','Tipo Brazalete', 'Cantidad Inicial', 'Cantidad Entregada', 'Inventario Final'])
    
    for item in queryset_brazaletes_acreditador:
        hoja_brazaletes_acreditador.append([item.nombre_acreditador, item.nombre_brazalete, item.cantidad_brazalete, item.cantidad_entregada, item.cantidad_resta])


    # Guardar el libro de Excel en la respuesta HTTP que lo mande el navegador
    wb.save(response)
    
    return response
@login_required
def listadoEventos(request):
    if not request.user.is_superuser:
        messages.error(request,'No cuenta con los permisos necesarios para acceder a esta sección.')
        return redirect('evento')
    eventos_cerrados = bkt_eventos.objects.filter(evento_activo=0, acreditacion_activa = 0).order_by('fecha_evento')
    return render(request,'listadoEventos.html',{'eventosCerrados':eventos_cerrados})
@login_required
def importarBrazaletes(request, id_evento):

    if not request.user.is_superuser:
        messages.error(request,'¡No tiene permisos para importar listados!')
        return redirect('evento')

    if request.method == 'POST' and 'archivo_excel_braza' in request.FILES:
        archivo = request.FILES['archivo_excel_braza']

        # Leer el archivo Excel utilizando pandas
        try:
            df = pd.read_excel(archivo, sheet_name='Hoja1', dtype={'CANTIDAD': str})
            df2 = pd.read_excel(archivo,sheet_name='Hoja2', dtype={'ACREDITADOR': str})
        except:
            # Manejar el error si el archivo no se puede leer correctamente
            messages.error(request, '¡El archivo no es válido!')
            return redirect('evento')

        # verifica que este iniciado el proceso de actualización

        iniciado_act = bkt_eventos.objects.get(id = id_evento)
        if iniciado_act.acreditacion_activa == 1:
            messages.error(request, '¡Ya se ha iniciado el proceso de acreditación!')
            return redirect('evento')      

        # Verificar las columnas requeridas
        columnas_requeridas = ['AREA', 'CANTIDAD']
        columnas_requeridas_braza = ['ACREDITADOR', 'AREA','CANTIDAD']

        columnas_excel = df.columns.tolist()
        columnas_excel_braza = df2.columns.tolist()

        if not set(columnas_requeridas).issubset(columnas_excel):
            # Manejar el error si alguna(s) columna(s) requerida(s) no está presente
            messages.error(
                request, '¡El archivo no contiene todas las columnas requeridas!')
            return redirect('evento')
        
        if not set(columnas_requeridas_braza).issubset(columnas_excel_braza):
            # Manejar el error si alguna(s) columna(s) requerida(s) no está presente
            messages.error(
                request, '¡El archivo no contiene todas las columnas requeridas aa!')
            return redirect('evento')

        # Validar campos vacíos
        registros = []
        for _, row in df.iterrows():
            if any(pd.isnull(row[columna]) or str(row[columna]).strip() == '' for columna in columnas_requeridas):
                # Manejar el error si hay campos vacíos en las columnas requeridas
                messages.error(
                    request, '¡El archivo contiene campos vacíos en las columnas requeridas! Por favor corrija e intente nuevamente.')
                return redirect('evento')
        
        for _, row in df2.iterrows():
            if any(pd.isnull(row[columna]) or str(row[columna]).strip() == '' for columna in columnas_requeridas):
                # Manejar el error si hay campos vacíos en las columnas requeridas
                messages.error(
                    request, '¡El archivo contiene campos vacíos en las columnas requeridas! Por favor corrija e intente nuevamente.')
                return redirect('evento')

        # Procesar los datos y guardar en la base de datos
        registros = []

        for _, row in df.iterrows():
            registro = inventarioBrazalete(
                id_evento=id_evento,
                nombre_brazalete=row['AREA'],
                cantidad_brazalete=row['CANTIDAD'],
                evento_cerrado = 0
            )
            registros.append(registro)
        
        registros_braza = []

        for _, row in df2.iterrows():
            registro = inventarioBrazaleteAcreditardor(
                id_evento=id_evento,
                nombre_acreditador=row['ACREDITADOR'],
                nombre_brazalete=row['AREA'],
                cantidad_brazalete=row['CANTIDAD'],
                evento_cerrado = 0
            )
            registros_braza.append(registro)

        
        if inventarioBrazalete.objects.filter(id_evento = id_evento).exists():
            inventarioBrazalete.objects.filter(id_evento = id_evento).delete()
        
        if inventarioBrazaleteAcreditardor.objects.filter(id_evento = id_evento).exists():
            inventarioBrazaleteAcreditardor.objects.filter(id_evento = id_evento).delete()

        inventarioBrazalete.objects.bulk_create(registros)
        inventarioBrazaleteAcreditardor.objects.bulk_create(registros_braza)
    
    messages.success(request, '¡Los datos se han importado exitosamente!')
    return redirect('evento')
@login_required
def verEstado(request, id_evento):

    if not request.user.is_superuser:
        messages.error(request,'No cuenta con los permisos necesarios para acceder a esta sección')
        return redirect('evento')
    
    evento_id = id_evento

    if not bkt_eventos.objects.filter(id = evento_id, evento_activo=1, acreditacion_activa = 1).exists():
        return redirect('evento')


    eventos_proceso = bkt_eventos.objects.filter(id = evento_id, evento_activo=1, acreditacion_activa = 1).order_by('fecha_evento')
    estado_brazalete = inventarioBrazalete.objects.filter(id_evento = evento_id)
    estado_brazalete_acreditador = inventarioBrazaleteAcreditardor.objects.filter(id_evento = evento_id).order_by('nombre_acreditador')
    total_color = acreditados_def.objects.filter(id_evento_id = evento_id).values('color_zona').annotate(total_registros_color = Count('color_zona'),
    acreditados=Count(Case (When(acreditado=True, then=1),ouput_field = IntegerField())),
    no_acreditados =Count(Case (When(acreditado=False, then=1),ouput_field = IntegerField())))

    #busca estadisticas
    total_acreditado = acreditados_def.objects.filter(id_evento_id = evento_id, acreditado = 1).count()
    total_registros = acreditados_def.objects.filter(id_evento_id = evento_id).count()
    porcentaje = round((total_acreditado /total_registros)*100,4)

    #grafico de totales
    total_brazalete = inventarioBrazalete.objects.filter(id_evento = evento_id)
    etiquetas = [brazalete.nombre_brazalete for brazalete in total_brazalete]
    valores = [cantidad.cantidad_brazalete for cantidad in total_brazalete]
    fig, ax = ptl.subplots()
    ax.pie(valores, labels=etiquetas,autopct='%1.1f%%',startangle=140)
    ax.axis('equal')
    ax.set_title('Distribución de brazaletes por Tipo')

    buffer = BytesIO()
    ptl.savefig(buffer, format='png')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.read()).decode()
    grafico1 = "data:image/png;base64," + image_base64

    #grafico acreditados x acreditador
    acreditados_x_acreditador = acreditados_def.objects.filter(id_evento=evento_id).values('acreditado_por').annotate(total_acreditado=Count('pk', filter=Q(acreditado=True)))

    etiquetas1 = [item['acreditado_por'] for item in acreditados_x_acreditador]
    valore1s = [item['total_acreditado'] for item in acreditados_x_acreditador]
    fig1, ax1 = ptl.subplots()
    ax1.pie(valore1s, labels=etiquetas1,autopct='%1.1f%%',startangle=140)
    ax1.axis('equal')
    ax1.set_title('Distribución de Acreditados por Acreditador')
    
    buffer1 = BytesIO()
    ptl.savefig(buffer1, format='png')
    buffer1.seek(0)
    image_base641 = base64.b64encode(buffer1.read()).decode()
    grafico2 = "data:image/png;base64," + image_base641 

    

    return render(request,'estadoEvento.html',{'eventoProceso':eventos_proceso, 'estado_brazalete':estado_brazalete,
                                                'estado_brazalete_acreditador':estado_brazalete_acreditador, 'total_acreditado':total_acreditado, 'total_registros':total_registros, 
                                                'porcentaje':porcentaje,'imagen':grafico1, 'imagen2':grafico2, 'total_color': total_color})
@login_required
def exportarPDFfinal(request, id):
    id_evento = id

    #DATOS DEL EVENTO#########

    datos_evento = bkt_eventos.objects.get(id = id_evento,evento_activo=0)
    nombre_evento = datos_evento.nombre_evento
    lugar_evento = datos_evento.lugar_evento
    fecha_evento = datos_evento.fecha_evento

    ###DATO TOTAL ACREDITABLES##
    total_acreditables = acreditados_def.objects.filter(id_evento_id = id_evento).count()
    total_acreditados = acreditados_def.objects.filter(id_evento_id = id_evento, acreditado = 1).count()
    total_faltante = total_acreditables - total_acreditados

    ###DATO ACREDITABLES POR ZONA##

    total_acreditables_zona = acreditados_def.objects.filter(id_evento_id=id_evento).values('zona_acceso').annotate(cantidad=Count('pk'))

    ###DATO ACREDITADOS POR ZONA##

    total_acreditados_zona = acreditados_def.objects.filter(id_evento_id=id_evento, acreditado = 1).values('zona_acceso').annotate(cantidad=Count('pk'))

    ###DATO ACREDITADOS POR ACREDITADOR##

    total_acreditados_acreditador = acreditados_def.objects.filter(id_evento_id=id_evento, acreditado = 1).values('acreditado_por').annotate(cantidad=Count('pk'))

    ###DATO INVENTARIO INICIAL DE BRAZALETES##

    inventario_brazaletes = inventarioBrazalete.objects.filter(id_evento = id_evento, evento_cerrado = 1)

    #grafico acreditados x acreditador
    acreditados_x_acreditador = acreditados_def.objects.filter(id_evento=id_evento).values('acreditado_por').annotate(total_acreditado=Count('pk', filter=Q(acreditado=True)))

    etiquetas1 = [item['acreditado_por'] for item in acreditados_x_acreditador]
    valore1s = [item['total_acreditado'] for item in acreditados_x_acreditador]
    fig1, ax1 = ptl.subplots()
    ax1.pie(valore1s, labels=etiquetas1,autopct='%1.1f%%',startangle=140)
    ax1.axis('equal')
    ax1.set_title('Distribución de Acreditados por Acreditador')
    
    buffer1 = BytesIO()
    ptl.savefig(buffer1, format='png')
    buffer1.seek(0)
    image_base641 = base64.b64encode(buffer1.read()).decode()
    grafico2 = "data:image/png;base64," + image_base641 


   
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    ancho_pagina, altura_pagina = letter = (21.59*cm, 27.94*cm)
    
    ####TITULO#########
    titulo = "REPORTE FINAL DE ACREDITACIÓN"
    ancho_texto = pdf.stringWidth(titulo, "Helvetica", 12)
    # Calcular la posición horizontal para centrar
    pos_x = (ancho_pagina - ancho_texto) / 2
    # Definir la posición vertical
    pos_y = altura_pagina - 2*cm

    
    

    # ESCRIBIR DATOS DEL PDF ####
    pdf.setFillColorRGB(0, 0, 1)
    pdf.drawString(pos_x, pos_y, titulo)
    pdf.setFillColorRGB(0, 0, 0)

    pdf.setFillColorRGB(1, 0, 0)
    pdf.drawString(2*cm, altura_pagina - 3.5*cm, "Evento: "+ nombre_evento)
    pdf.drawString(2*cm, altura_pagina - 4*cm, "Lugar: "+ lugar_evento)
    pdf.drawString(2*cm, altura_pagina - 4.5*cm, "Fecha: "+ str(fecha_evento))
    pdf.setFillColorRGB(0, 0, 0)

    pdf.drawString(2*cm, altura_pagina - 5.5*cm, "Total Acreditables: "+ str(total_acreditables))
    pdf.drawString(2*cm, altura_pagina - 6*cm, "Total Acreditados: "+ str(total_acreditados))
    pdf.drawString(2*cm, altura_pagina - 6.5*cm, "Total Faltantes: "+ str(total_faltante))
    
    pdf.setFillColorRGB(0, 0, 1)
    
    pdf.drawString(2*cm, altura_pagina - 7.5*cm, "Total Acreditables por Zona:")

    pdf.setFillColorRGB(0, 0, 0)
    x = 8.5*cm
    for zona in total_acreditables_zona:
        acreditables_zona = f"Zona: {zona['zona_acceso']}    -    Cantidad: {zona['cantidad']}"
        pdf.drawString(2*cm, altura_pagina - x, acreditables_zona)
        x += 0.5*cm
    
    x += 1*cm
    pdf.setFillColorRGB(0, 0, 1)
    pdf.drawString(2*cm, altura_pagina - x, "Total Acreditados por Zona:")
    pdf.setFillColorRGB(0, 0, 0)
    x += 0.5*cm
    for zona in total_acreditados_zona:
        acreditados_zona = f"Zona: {zona['zona_acceso']}    -    Cantidad: {zona['cantidad']}"
        pdf.drawString(2*cm, altura_pagina - x, acreditados_zona)
        x += 0.5*cm

    x += 0.5*cm
    pdf.setFillColorRGB(0, 0, 1)
    pdf.drawString(2*cm, altura_pagina - x, "Total Acreditados por Acreditador:")
    pdf.setFillColorRGB(0, 0, 0)
    x += 0.5*cm
    for acreditador in total_acreditados_acreditador:
        acreditados_acredit = f"Acreditador: {acreditador['acreditado_por']}    -    Cantidad: {acreditador['cantidad']}"
        pdf.drawString(2*cm, altura_pagina - x, acreditados_acredit)
        x += 0.5*cm

    x += 0.5*cm
    pdf.setFillColorRGB(0, 0, 1)
    pdf.drawString(2*cm, altura_pagina - x, "Inventario Inicial de Brazaletes:")
    pdf.setFillColorRGB(0, 0, 0)
    x += 0.5*cm
    for brazalete in inventario_brazaletes:
        total_por_zona = f'Zona: {brazalete.nombre_brazalete}    -    Cantidad: {brazalete.cantidad_brazalete}'
        pdf.drawString(2*cm, altura_pagina - x, total_por_zona)
        x += 0.5*cm
        
    x += 0.5*cm
    pdf.setFillColorRGB(0, 0, 1)
    pdf.drawString(2*cm, altura_pagina - x, "Inventario Final de Brazaletes:")
    pdf.setFillColorRGB(0, 0, 0)
    x += 0.5*cm
    for brazalete_final in inventario_brazaletes:
        total_restante = brazalete_final.cantidad_brazalete - brazalete_final.cantidad_entregada
        total_por_zona_restante = f'Zona: {brazalete_final.nombre_brazalete}    - Cantidad entregada: {brazalete_final.cantidad_entregada} - Sobrante: {total_restante}'
        pdf.drawString(2*cm, altura_pagina - x, total_por_zona_restante)
        x += 0.5*cm
    
    x += 0.5*cm
    
    pdf.showPage() #final 

    # pdf.drawImage(grafico2,2*cm, altura_pagina - 3.5*cm, width=6*cm, height=5*cm)

    # pdf.showPage()

    pdf.save()

    buffer.seek(0)

    nombre_archivo = nombre_evento+'_'+str(fecha_evento)+'.pdf'

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename= {nombre_archivo}'


    return response

def importaAdicionales(request, id_evento):

    if not request.user.is_superuser:
        messages.error(request,'¡No tiene permisos para importar listados!')
        return redirect('evento')

    if request.method == 'POST' and 'archivo_excel' in request.FILES:
        archivo = request.FILES['archivo_excel']

        # Leer el archivo Excel utilizando pandas
        try:
            df = pd.read_excel(archivo, dtype={'NUMERO_DOCUMENTO': str})
        except:
            # Manejar el error si el archivo no se puede leer correctamente
            messages.error(request, '¡El archivo no es válido!')
            return redirect('evento')

        # Verificar las columnas requeridas
        columnas_requeridas = ['NOMBRES', 'APELLIDOS', 'TIPO_DOCUMENTO',
                               'NUMERO_DOCUMENTO', 'CARGO Y O FUNCIÓN', 'AREA_DE_TRABAJO', 'COLOR_ZONA_BRAZALETE']
        columnas_excel = df.columns.tolist()
        if not set(columnas_requeridas).issubset(columnas_excel):
            # Manejar el error si alguna(s) columna(s) requerida(s) no está presente
            messages.error(
                request, '¡El archivo no contiene todas las columnas requeridas!')
            return redirect('evento')

        # Validar campos vacíos
        registros = []
        for _, row in df.iterrows():
            if any(pd.isnull(row[columna]) or str(row[columna]).strip() == '' for columna in columnas_requeridas):
                # Manejar el error si hay campos vacíos en las columnas requeridas
                messages.error(
                    request, '¡El archivo contiene campos vacíos en las columnas requeridas! Por favor corrija e intente nuevamente.')
                return redirect('evento')

        # Procesar los datos y guardar en la base de datos
        registros = []

        for  _, row in df.iterrows():
            registro = acreditados_tmp(
                nombre_persona=row['NOMBRES'],
                apellido_persona=row['APELLIDOS'],
                tipo_doc=row['TIPO_DOCUMENTO'],
                numero_doc=row['NUMERO_DOCUMENTO'],
                cargo=row['CARGO Y O FUNCIÓN'],
                zona_acceso=row['AREA_DE_TRABAJO'],
                color_zona=row['COLOR_ZONA_BRAZALETE'],
                empresa=row['EMPRESA'],
                id_evento_id = id_evento
            )
            registros.append(registro)

        acreditados_tmp.objects.bulk_create(registros)

        tabla_origen = acreditados_tmp.objects.filter(id_evento_id = id_evento)

        tabla_destino =[]  
            
        for dato in tabla_origen:
            registros = acreditados_def(
            nombre_persona = dato.nombre_persona,
            apellido_persona = dato.apellido_persona,
            tipo_doc = dato.tipo_doc,
            numero_doc = dato.numero_doc,
            cargo = dato.cargo,
            zona_acceso = dato.zona_acceso,
            color_zona = dato.color_zona,
            empresa = dato.empresa,
            id_evento_id = id_evento)

            # validar que no este duplicado
            if not acreditados_def.objects.filter(numero_doc = dato.numero_doc, id_evento = id_evento).exists(): 
                tabla_destino.append(registros)

        acreditados_def.objects.bulk_create(tabla_destino)

        acreditados_tmp.objects.filter(id_evento_id=id_evento).delete()

        messages.success(request, '¡Los datos se han importado exitosamente!')
        return redirect('evento')

    return redirect('evento')           
        